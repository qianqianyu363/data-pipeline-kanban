#!/usr/bin/env python3
"""
看板数据录入模板-调整.xlsx → data.json 解析脚本
读取新模板4个Sheet + 问题反馈，生成看板可用的 data.json
"""
import openpyxl, json, sys, os, re, datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')


def sync_embed_data(index_html_path, data):
    """把解析结果同步写入 index.html 的 dataEmbed 内嵌块（离线 fallback 用，保持数据新鲜）。
    仅在 index.html 存在且含 dataEmbed 块时生效，幂等。"""
    if not os.path.exists(index_html_path):
        print(f"   ⚠️ index.html 不存在，跳过内嵌数据同步: {index_html_path}")
        return False
    with open(index_html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    pat = re.compile(r'(<script id="dataEmbed" type="application/json">)([\s\S]*?)(</script>)')
    if not pat.search(html):
        print("   ⚠️ index.html 未找到 dataEmbed 块，跳过内嵌数据同步")
        return False
    new_json = json.dumps(data, ensure_ascii=False, indent=2)
    html_new = pat.sub(lambda m: m.group(1) + '\n' + new_json + '\n' + m.group(3), html, count=1)
    with open(index_html_path, 'w', encoding='utf-8') as f:
        f.write(html_new)
    print(f"   ✅ 已同步内嵌数据到 index.html (dataEmbed 块, {len(new_json)} 字符)")
    return True

# 支持命令行指定模板/输出：python parse_template.py [模板.xlsx] [输出.json]
_default_tpl = '看板数据录入模板-调整.xlsx'
_default_out = 'data.json'
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), sys.argv[1] if len(sys.argv) > 1 else _default_tpl)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), sys.argv[2] if len(sys.argv) > 2 else _default_out)
LEDGER_PATH = os.path.join(os.path.dirname(__file__), '需求下发登记.json')  # 对话登记台账

# 注意：用 data_only=False 保留公式文本，配合下方公式求值器自算派生值。
# openpyxl 保存会丢公式缓存（升级版模板 data_only=True 读出来全是 None），故不依赖缓存。
wb = openpyxl.load_workbook(TEMPLATE_PATH)


def colnum(letter):
    """列字母 → 数字（A=1, B=2 …）"""
    n = 0
    for ch in str(letter).upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def eval_formula(formula, row, get_val, seen=None):
    """求值模板公式（=A{r}*K、=A-B、=A/B 等算术，列引用限同行）。
    非公式值原样返回；公式求值失败返回 None。"""
    if seen is None:
        seen = set()
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    expr = formula[1:]
    def repl(m):
        col, r = m.group(1).replace('$', ''), int(m.group(2))
        key = (col.upper(), r)
        if key in seen:  # 防循环引用
            return '0'
        seen.add(key)
        v = get_val(col.upper(), r)
        if isinstance(v, str) and v.startswith('='):
            v = eval_formula(v, row, get_val, seen)
        try:
            return str(float(v))
        except (TypeError, ValueError):
            return '0'
    expr = re.sub(r'(\$?[A-Za-z]{1,3}\$?)(\d+)', repl, expr)
    try:
        return float(eval(expr, {'__builtins__': {}}, {}))
    except Exception:
        return None


def cell_val(ws, col, r):
    """取单元格值：公式单元格递归求值，非公式返回原值。"""
    if isinstance(col, int):
        from openpyxl.utils import get_column_letter
        col = get_column_letter(col)
    v = ws.cell(row=r, column=colnum(col)).value
    if isinstance(v, str) and v.startswith('='):
        return eval_formula(v, r, lambda c, rr: cell_val(ws, c, rr))
    return v

# ============================================================
# 1. 各项需求跟踪 → tasks
# ============================================================
ws1 = wb['各项需求跟踪']
tasks = []
month_names_set = set()

# 列映射 (列号→字段名) — 适配升级版模板（各项需求跟踪 27 列）
COL_MAP = {
    2: 'month_raw',    # 月份
    3: 'demandItem',   # 需求项
    4: 'settleType',   # 结算类型
    5: 'name',         # 需求名称
    6: 'source',       # 需求来源
    7: 'dueMonth',     # 约定采集完成时间（新）
    9: 'matQty',       # 材料数量（新，需求准确率分母）
    12: 'timelyRate',  # 需求发布及时率 (0/1)
    13: 'totalPoints', # 下发数量(含多家）
    14: 'itemCount',   # 下发材料
    15: 'reviewIssues',# 预审问题数据
    18: 'accuracyRate',# 需求准确率（自算 = itemCount/matQty）
    22: 'collected',   # 实际采集完成量
    24: 'completionRate', # 采集完成率（自算 = collected/totalPoints）
    25: 'invalidData', # 无效数据
    26: 'invalidRate', # 无效数据占比（自算 = invalidData/collected）
    27: 'published',   # 发布数据
}

# 月份格式化
def fmt_month(m):
    m = str(m or '').strip()
    if '月' in m:
        num = m.replace('月', '').strip()
        if num.isdigit():
            return f"{int(num):02d}月"
        return m
    if m.isdigit():
        return f"{int(m):02d}月"
    return m


def fmt_due_month(v):
    """约定采集完成时间 → 'MM月'。兼容 datetime/日期序列号/文本('7月'、'2026-07-31')"""
    if v is None or v == '':
        return None
    if isinstance(v, datetime.datetime):
        return f"{v.month:02d}月"
    if isinstance(v, datetime.date):
        return f"{v.month:02d}月"
    if isinstance(v, (int, float)):
        try:  # Excel 日期序列号（1900-01-01 基准，实际用 1899-12-30）
            dt = datetime.date(1899, 12, 30) + datetime.timedelta(days=float(v))
            return f"{dt.month:02d}月"
        except Exception:
            return None
    s = str(v).strip()
    if '月' in s:  # '7月' / '2026年7月'
        digits = ''.join(c for c in s.replace('年', '') if c.isdigit())
        return f"{int(digits):02d}月" if digits else None
    for sep in ('-', '.', '/'):  # '2026-07-31'
        if sep in s:
            parts = s.split(sep)
            if len(parts) >= 2 and parts[1].strip().isdigit():
                return f"{int(parts[1].strip()):02d}月"
    return None


# ============================================================
# 1.5 对话登记台账合并
# ============================================================
def load_ledger(path=LEDGER_PATH):
    """读取对话登记台账 records。文件不存在/空/坏 JSON/结构无效均容错返回 []。"""
    if not os.path.exists(path):
        print("   ℹ️ 台账文件不存在，跳过合并:", path)
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            raw = f.read().strip()
        if not raw:
            print("   ℹ️ 台账文件为空，跳过合并")
            return []
        obj = json.loads(raw)
    except Exception as e:
        print("   ⚠️ 台账 JSON 解析失败，跳过合并:", e)
        return []
    if not isinstance(obj, dict) or not isinstance(obj.get('records'), list):
        print("   ⚠️ 台账结构无效（需 {records:[...]}），跳过合并")
        return []
    return obj['records']


def _norm_num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_int(v):
    return int(v) if v is not None else None


def _to_pct(v):
    return round(v, 1) if v is not None else None


def merge_ledger_tasks(tasks, records, month_names_set):
    """把台账 records 并入 tasks（去重 name+month、字段归一化、缺失补 None、派生值自算）。
    模板优先：与模板任务重复的台账记录跳过，不覆盖。返回新增条数。
    副作用：修改 tasks（append）与 month_names_set（补新月份）。"""
    existing_keys = {(t.get('name'), t.get('month')) for t in tasks}
    seen_in_ledger = set()
    merged = skipped_dup = skipped_tpl = 0

    for rec in records:
        if not isinstance(rec, dict):
            continue
        name = str(rec.get('name') or '').strip()
        month = fmt_month(rec.get('month'))
        if not name or not month or '月' not in month:
            print(f"   ⚠️ 台账记录缺 name/month，跳过: {name!r}/{month!r}")
            continue
        key = (name, month)

        # 台账内部去重（只保留首条）
        if key in seen_in_ledger:
            skipped_dup += 1
            print(f"   ⚠️ 台账内部重复，跳过: {name}/{month}")
            continue
        seen_in_ledger.add(key)

        # 与模板 tasks 去重（模板优先，不覆盖模板）
        if key in existing_keys:
            skipped_tpl += 1
            print(f"   ⚠️ 台账与模板重复，跳过（以模板为准）: {name}/{month}")
            continue

        # ---- 字段读取与派生自算 ----
        item_cnt = _norm_num(rec.get('itemCount'))
        req_qty = _norm_num(rec.get('reqQty'))
        mat_qty = _norm_num(rec.get('matQty'))
        total_pts = _norm_num(rec.get('totalPoints'))
        if total_pts is None and item_cnt is not None and req_qty is not None:
            total_pts = item_cnt * req_qty          # 下发数量 = 下发材料 × 要求数量
        if item_cnt is None and total_pts is not None and req_qty:
            item_cnt = total_pts / req_qty          # 反向补下发材料

        collected = _norm_num(rec.get('collected'))
        comp_rate = _norm_num(rec.get('completionRate'))
        if comp_rate is None and total_pts and collected is not None:
            comp_rate = collected / total_pts * 100  # 采集完成率 = 采集量/下发量

        accuracy = _norm_num(rec.get('accuracyRate'))
        if accuracy is None and item_cnt is not None and mat_qty:
            accuracy = item_cnt / mat_qty * 100      # 需求准确率 = 下发材料/材料数量

        published = _norm_num(rec.get('publishedData'))
        timely = _norm_num(rec.get('timelyRate'))

        task = {
            "name": name,
            "settleType": rec.get('settleType') or '价格采集',
            "source": rec.get('source') or None,
            "demandItem": rec.get('demandItem') or '市场价',
            "month": month,
            "dueMonth": fmt_due_month(rec.get('dueMonth')) if rec.get('dueMonth') else None,
            "totalPoints": _to_int(total_pts),
            "itemCount": _to_int(item_cnt),
            "reviewIssues": _to_int(_norm_num(rec.get('reviewIssues'))),
            "issuedQty": _to_int(total_pts),
            "collected": _to_int(collected),
            "completionRate": _to_pct(comp_rate),
            "accuracyRate": _to_pct(accuracy),
            "invalidData": _to_int(_norm_num(rec.get('invalidData'))),
            "invalidRate": _to_pct(_norm_num(rec.get('invalidRate'))),
            "publishedData": _to_int(published),
            "timelyRate": _to_int(timely),
            "feedbackMat": _norm_num(rec.get('feedbackMat')),
            "expertVerify": _norm_num(rec.get('expertVerify')),
            "supplierRejects": _norm_num(rec.get('supplierRejects')),
            "closed": rec.get('closed') or None,
            "allClosed": rec.get('allClosed') or None,
            "notes": rec.get('notes') or None,
            "reqQty": _to_int(req_qty),   # 仅台账来源任务携带，前端忽略
        }
        tasks.append(task)
        if month:
            month_names_set.add(month)    # ★ 让 months/filterOptions/monthly_trend 自动含新月份
        existing_keys.add(key)
        merged += 1

    print(f"   📋 台账合并: 新增 {merged} 条 | 台账内部重复跳过 {skipped_dup} | 与模板重复跳过 {skipped_tpl}")
    return merged


for r in range(3, ws1.max_row + 1):
    name = ws1.cell(row=r, column=5).value
    if not name:
        continue  # 跳过空行

    month_raw = ws1.cell(row=r, column=2).value
    month = fmt_month(month_raw)
    month_names_set.add(month)

    demand_item = ws1.cell(row=r, column=3).value or ''
    settle_type = ws1.cell(row=r, column=4).value or ''
    source = ws1.cell(row=r, column=6).value or ''

    def nv(col):
        v = cell_val(ws1, col, r)   # 公式单元格自动求值
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    total_pts = nv(13)
    item_cnt = nv(14)
    mat_qty = nv(9)        # 材料数量（需求准确率分母）
    review_issues = nv(15)
    collected = nv(22)
    comp_rate = nv(24)     # 模板缓存（升级版 openpyxl 保存后无缓存，通常 None）
    accuracy = nv(18)      # 模板缓存
    invalid = nv(25)
    invalid_rate = nv(26)  # 模板缓存
    published = nv(27)
    timely = nv(12)
    due_month_raw = ws1.cell(row=r, column=7).value  # 约定采集完成时间

    # ---- 派生值自算（openpyxl 保存丢公式缓存，必须自算，遵循模板公式语义）----
    # 采集完成率 = 实际采集完成量 / 下发数量(含多家)
    if comp_rate is None and total_pts:
        comp_rate = collected / total_pts if collected is not None else None
    # 需求准确率 = 下发材料 / 材料数量（模板公式 =N/I 语义，非 1-预审问题占比）
    if accuracy is None and mat_qty:
        accuracy = item_cnt / mat_qty if item_cnt is not None else None
    # 无效数据占比 = 无效数据 / 实际采集完成量
    if invalid_rate is None and collected:
        invalid_rate = invalid / collected if invalid is not None else None
    due_month = fmt_due_month(due_month_raw)

    task = {
        "name": str(name).strip(),
        "settleType": settle_type,
        "source": source if source else None,
        "demandItem": demand_item,
        "month": month,
        "dueMonth": due_month,   # 约定采集完成时间（新，用于当月采集完成率）
        "totalPoints": int(total_pts) if total_pts is not None else None,
        "itemCount": int(item_cnt) if item_cnt is not None else None,
        "reviewIssues": int(review_issues) if review_issues is not None else None,
        "issuedQty": int(total_pts) if total_pts is not None else None,
        "collected": int(collected) if collected is not None else None,
        "completionRate": round(comp_rate * 100, 1) if comp_rate is not None and comp_rate != '#DIV/0!' else None,
        "accuracyRate": round(accuracy * 100, 1) if accuracy is not None and isinstance(accuracy, (int, float)) and accuracy != '#DIV/0!' else None,
        "invalidData": int(invalid) if invalid is not None else None,
        "invalidRate": round(invalid_rate * 100, 1) if invalid_rate is not None else None,
        "publishedData": int(published) if published is not None else None,
        "timelyRate": int(timely) if timely is not None else None,
        # 缺失字段置空
        "feedbackMat": None,
        "expertVerify": None,
        "supplierRejects": None,
        "closed": None,
        "allClosed": None,
        "notes": None,
    }
    tasks.append(task)

# ---- 合并对话登记台账（需求下发登记.json）----
_ledger_records = load_ledger()
_merged_count = merge_ledger_tasks(tasks, _ledger_records, month_names_set)

# 按月份排序
month_order = {f"{i:02d}月": i for i in range(1, 13)}
tasks.sort(key=lambda t: (month_order.get(t['month'], 99), t['name'] or ''))

# ============================================================
# 2. 预算与结算 → settlements
# ============================================================
ws2 = wb['预算与结算']
settlements = []
for r in range(3, ws2.max_row + 1):
    month = ws2.cell(row=r, column=1).value
    module = ws2.cell(row=r, column=2).value
    demand_type = ws2.cell(row=r, column=3).value
    qty = cell_val(ws2, 4, r)
    price = cell_val(ws2, 5, r)
    # 去年值5列（新，col7-11）
    last_module = ws2.cell(row=r, column=7).value
    last_type = ws2.cell(row=r, column=8).value
    last_qty = cell_val(ws2, 9, r)
    last_price = cell_val(ws2, 10, r)

    if not month or not module or not demand_type:
        continue

    def to_num(v):
        if v is None: return 0
        try: return float(v)
        except: return 0

    q = to_num(qty); p = to_num(price)
    lq = to_num(last_qty); lp = to_num(last_price)
    has_last = last_qty is not None or last_price is not None

    settlements.append({
        "月份": str(month).strip(),
        "产品模块": str(module).strip(),
        "需求类型": str(demand_type).strip(),
        "需求数量": int(q),
        "结算单价": p,
        # 结算金额 = 需求数量 × 结算单价（openpyxl 丢缓存，自算）
        "结算金额": round(q * p, 2),
        # 去年对比（新，供前端"费用执行去年对比"）
        "去年产品模块": str(last_module).strip() if last_module else None,
        "去年需求类型": str(last_type).strip() if last_type else None,
        "去年需求数量": int(lq) if has_last else None,
        "去年结算单价": round(lp, 2) if last_price is not None else None,
        # 去年支出费用 = 去年需求数量 × 去年结算单价
        "去年支出费用": round(lq * lp, 2) if last_price is not None else None,
    })

# ============================================================
# 3. 质量抽审 → qualityReview (by_month 月度格式)
# ============================================================
ws3 = wb['质量抽审']
quality_review = {}

current_task = None
current_month = None
for r in range(1, ws3.max_row + 1):
    a = ws3.cell(row=r, column=1).value  # 月份
    b = ws3.cell(row=r, column=2).value  # 抽审任务
    c = ws3.cell(row=r, column=3).value  # 维度
    d = ws3.cell(row=r, column=4).value  # 项
    e = ws3.cell(row=r, column=5).value  # 值

    if a and b:
        current_task = str(b).strip()
        current_month = str(a).strip()
        if current_task not in quality_review:
            quality_review[current_task] = {"by_month": {}}
        if current_month not in quality_review[current_task]["by_month"]:
            quality_review[current_task]["by_month"][current_month] = {}

    if current_task and current_month and d is not None and e is not None:
        quality_review[current_task]["by_month"][current_month][str(d).strip()] = e

# ============================================================
# 4. 加工预审 → preReview (统计摘要)
# ============================================================
ws4 = wb['加工预审']
preview_stats = {
    "total": 0,
    "passed": 0,
    "failed": 0,
    "by_reviewer": defaultdict(lambda: {"total": 0, "passed": 0}),
    "by_month": defaultdict(lambda: {"total": 0, "passed": 0}),
    "fail_reasons": defaultdict(int),
    "new_enterprise": 0,
    "old_enterprise": 0,                       # 老供应商（更新登记表）
    "price_audit": {"total": 0, "passed": 0},  # 价格审核（新，col9）
}

has_data = False
for r in range(3, ws4.max_row + 1):
    reviewer = ws4.cell(row=r, column=3).value
    passed = ws4.cell(row=r, column=8).value
    month = ws4.cell(row=r, column=2).value
    reason = ws4.cell(row=r, column=10).value      # 升级后"不通过原因"顺延 col10
    is_new = ws4.cell(row=r, column=7).value
    price_audit = ws4.cell(row=r, column=9).value  # 新增"价格审核结果"

    if not reviewer:
        continue
    has_data = True

    preview_stats["total"] += 1
    passed_flag = str(passed or '').strip() == '是'
    if passed_flag:
        preview_stats["passed"] += 1
    else:
        preview_stats["failed"] += 1

    reviewer_name = str(reviewer).strip()
    preview_stats["by_reviewer"][reviewer_name]["total"] += 1
    if passed_flag:
        preview_stats["by_reviewer"][reviewer_name]["passed"] += 1

    m = str(month or '').strip()
    if m:
        preview_stats["by_month"][m]["total"] += 1
        if passed_flag:
            preview_stats["by_month"][m]["passed"] += 1

    if reason and not passed_flag:
        rsn = str(reason).strip()
        if rsn:
            preview_stats["fail_reasons"][rsn] += 1

    # 新老供应商占比
    new_flag = str(is_new or '').strip()
    if new_flag in ('是', '新增'):
        preview_stats["new_enterprise"] += 1
    elif new_flag in ('更新登记表', '否'):
        preview_stats["old_enterprise"] += 1
    # 价格审核通过统计（col9）
    pa = str(price_audit or '').strip()
    if pa in ('是', '通过'):
        preview_stats["price_audit"]["total"] += 1
        preview_stats["price_audit"]["passed"] += 1
    elif pa in ('否', '不通过'):
        preview_stats["price_audit"]["total"] += 1

# 转换defaultdict为普通dict
if has_data:
    preview_stats["by_reviewer"] = dict(preview_stats["by_reviewer"])
    preview_stats["by_month"] = dict(preview_stats["by_month"])
    preview_stats["fail_reasons"] = dict(preview_stats["fail_reasons"])
    # 计算通过率
    total = preview_stats["total"]
    preview_stats["passRate"] = round(preview_stats["passed"] / total * 100, 1) if total else 0
    pa_total = preview_stats["price_audit"]["total"]
    preview_stats["price_audit"]["passRate"] = round(preview_stats["price_audit"]["passed"] / pa_total * 100, 1) if pa_total else 0

    for rv in preview_stats["by_reviewer"].values():
        rv["passRate"] = round(rv["passed"] / rv["total"] * 100, 1) if rv["total"] else 0
    for bm in preview_stats["by_month"].values():
        bm["passRate"] = round(bm["passed"] / bm["total"] * 100, 1) if bm["total"] else 0
else:
    preview_stats = None

# ============================================================
# 5. 问题反馈 → issueFeedback
# ============================================================
ws5 = wb['问题反馈']
issue_feedback = {}
current_section = "main"
issue_labels = {}

for r in range(1, ws5.max_row + 1):
    label = ws5.cell(row=r, column=2).value
    val = ws5.cell(row=r, column=3).value

    if not label:
        continue
    label = str(label).strip()

    if label == "已共识问题标签":
        current_section = "labels"
        continue

    if current_section == "main":
        try:
            issue_feedback[label] = int(float(val)) if val is not None else None
        except (ValueError, TypeError):
            issue_feedback[label] = val
    else:
        try:
            issue_labels[label] = int(float(val)) if val is not None else None
        except (ValueError, TypeError):
            issue_labels[label] = val

if issue_labels:
    issue_feedback["已共识问题标签"] = issue_labels

# ============================================================
# 6. 产品验收 → productAcceptance
# ============================================================
product_acceptance = {"by_month": {}}
if '产品验收' in wb.sheetnames:
    ws6 = wb['产品验收']
    for r in range(2, ws6.max_row + 1):
        month = ws6.cell(row=r, column=1).value
        label = ws6.cell(row=r, column=2).value
        val = ws6.cell(row=r, column=3).value
        if not month or not label:
            continue
        m = str(month).strip()
        if m == '月份':
            continue
        if m not in product_acceptance["by_month"]:
            product_acceptance["by_month"][m] = {}
        try:
            product_acceptance["by_month"][m][str(label).strip()] = float(val) if val else val
        except (ValueError, TypeError):
            product_acceptance["by_month"][m][str(label).strip()] = val

# ============================================================
# 7. 采集产能 → collectionCapacity（新 Sheet）
# ============================================================
collection_capacity = []
if '采集产能' in wb.sheetnames:
    ws7 = wb['采集产能']
    for r in range(2, ws7.max_row + 1):
        month = ws7.cell(row=r, column=1).value
        if not month:
            continue

        def cap_num(v):
            if v is None: return None
            try: return float(v)
            except: return None

        collection_capacity.append({
            "月份": str(month).strip(),
            "计划采集家数": cap_num(ws7.cell(row=r, column=2).value),
            "计划采集条数": cap_num(ws7.cell(row=r, column=3).value),
            "实际采集家数": cap_num(ws7.cell(row=r, column=4).value),
            "实际采集条数": cap_num(ws7.cell(row=r, column=5).value),
            "员工总人数": cap_num(ws7.cell(row=r, column=6).value),
            "备注": ws7.cell(row=r, column=7).value,
        })

# ============================================================
# 8. 计算月度趋势 monthlyTrend
# ============================================================
monthly_trend = defaultdict(list)
months_in_data = sorted(set(
    f"{int(m.split('月')[0]):02d}月" for m in month_names_set
    if m and '月' in m and m.split('月')[0].isdigit()
), key=lambda x: int(x.split('月')[0]))

for m in months_in_data:
    month_tasks = [t for t in tasks if t['month'] == m]
    total_pts = sum(t['totalPoints'] or 0 for t in month_tasks)
    total_issued = sum(t['issuedQty'] or 0 for t in month_tasks)
    total_collected = sum(t['collected'] or 0 for t in month_tasks)
    total_review = sum(t['reviewIssues'] or 0 for t in month_tasks)
    total_invalid = sum(t['invalidData'] or 0 for t in month_tasks)

    # 采集完成率
    comp_rate = round(total_collected / total_issued * 100, 1) if total_issued else 0
    # 需求准确率
    acc_rate = round((1 - total_review / total_pts) * 100, 1) if total_pts else 0
    # 无效数据率
    inv_rate = round(total_invalid / total_collected * 100, 1) if total_collected else 0

    monthly_trend["需求总条数"].append(total_pts)
    monthly_trend["采集下发数"].append(total_issued)
    monthly_trend["采集完成数"].append(total_collected)
    monthly_trend["采集完成率"].append(comp_rate)
    monthly_trend["需求准确率"].append(acc_rate)
    monthly_trend["无效数据数"].append(total_invalid)
    monthly_trend["无效数据率"].append(inv_rate)

# ============================================================
# 9. 构建完整 data.json
# ============================================================
# 收集唯一值
demand_items = sorted(set(t['demandItem'] for t in tasks if t['demandItem']))
settle_types = sorted(set(t['settleType'] for t in tasks if t['settleType']))
sources = sorted(set(t['source'] for t in tasks if t['source']))

# 去重月份（台账合并后 months_in_data 已含台账月份，去重避免与硬编码 07/08 重复）
month_keys = list(dict.fromkeys(list(months_in_data) + ["07月", "08月"]))
month_names = {m.replace('月', '').zfill(2): m for m in months_in_data}

# 构建monthlyTrend的key（按月）
monthly_trend_dict = dict(monthly_trend)

# 额外：从结算数据算预算执行率和成本执行率
total_budget = 3900000
total_cost = sum(s['结算金额'] for s in settlements)
budget_rate = round(total_cost / total_budget * 100, 1) if total_budget else 0

# 为monthlyTrend补充预算执行率和成本执行率
# 按月份计算预算执行率
for m_idx, m_name in enumerate(months_in_data):
    m_num = m_name.replace('月', '')
    month_settlements = [s for s in settlements if s['月份'].replace('月', '') == m_num or s['月份'] == m_num + '月']
    month_cost = sum(s['结算金额'] for s in month_settlements)
    month_budget_rate = round(month_cost / (total_budget / 6) * 100, 1) if (total_budget / 6) else 0

    if '预算执行率' not in monthly_trend_dict:
        monthly_trend_dict['预算执行率'] = []
    if '成本执行率' not in monthly_trend_dict:
        monthly_trend_dict['成本执行率'] = []

    # 累计执行率
    cumulative_cost = sum(s['结算金额'] for s in settlements
                         if int(s['月份'].replace('月', '')) <= int(m_num))
    cum_budget_rate = round(cumulative_cost / total_budget * 100, 1) if total_budget else 0

    if len(monthly_trend_dict['预算执行率']) < len(months_in_data):
        monthly_trend_dict['预算执行率'].append(cum_budget_rate)
    if len(monthly_trend_dict['成本执行率']) < len(months_in_data):
        monthly_trend_dict['成本执行率'].append(cum_budget_rate)

data = {
    "title": "需求下发登记全链路看板",
    "updateDate": datetime.date.today().isoformat(),
    "months": month_keys,
    "monthNames": month_names,
    "filterOptions": {
        "demandItems": ["全部"] + demand_items,
        "settleTypes": ["全部"] + settle_types,
        "sources": ["全部"] + sources,
        "months": ["全部"] + month_keys,
    },
    "tasks": tasks,
    "settlements": settlements,
    "totalBudget": total_budget,
    "stepConfig": {
        "需求下发": {
            "color": "#E74C3C",
            "icon": "📋",
            "metrics": [
                {"key": "totalPoints", "label": "需求总条数", "unit": "条"},
                {"key": "itemCount", "label": "需求项数", "unit": "项"},
                {"key": "reviewIssues", "label": "初审问题数", "unit": "条"},
                {"key": "reviewIssueRate", "label": "初审问题率", "unit": "%"},
                {"key": "reviewPassRate", "label": "初审通过率", "unit": "%"},
                {"key": "timelyRate", "label": "发布及时率", "unit": "%"},
            ]
        },
        "采集执行": {
            "color": "#F39C12",
            "icon": "🔧",
            "metrics": [
                {"key": "issuedQty", "label": "采集下发数", "unit": "条"},
                {"key": "collected", "label": "采集完成数", "unit": "条"},
                {"key": "completionRate", "label": "采集完成率", "unit": "%"},
                {"key": "invalidData", "label": "无效数据", "unit": "条"},
                {"key": "publishedData", "label": "发布数据", "unit": "条"},
            ]
        },
        "数据加工/审核": {
            "color": "#F1C40F",
            "icon": "🔍",
            "metrics": [
                {"key": "accuracyRate", "label": "需求准确率", "unit": "%"},
                {"key": "feedbackMat", "label": "材料反馈数", "unit": "条"},
                {"key": "expertVerify", "label": "专家核实通过数", "unit": "条"},
            ]
        },
        "上线发布": {
            "color": "#2ECC71",
            "icon": "🚀",
            "metrics": [
                {"key": "closedRate", "label": "闭环率", "unit": "%"},
                {"key": "allClosedRate", "label": "全部闭环率", "unit": "%"},
            ]
        },
    },
    "specialProjects": {
        "采购价更新": {
            "color": "#9B59B6",
            "icon": "🟣",
            "desc": "月度大批量更新需求专项",
            "metrics": [
                {"key": "updateDemand", "label": "更新需求量", "unit": "条"},
                {"key": "updateCompletion", "label": "更新完成率", "unit": "%"},
            ]
        },
        "结算与成本": {
            "color": "#1ABC9C",
            "icon": "💰",
            "desc": "外包、专家人力成本预算管控",
            "metrics": [
                {"key": "budgetRate", "label": "预算执行率", "unit": "%"},
                {"key": "totalCost", "label": "累计成本", "unit": "万"},
            ]
        },
    },
    "alerts": [
        {
            "condition": "completionRate<60",
            "priority": "P0",
            "problem": "{name} 采集完成率 {value}",
            "action": "专项攻坚，增配采集资源",
        },
        {
            "condition": "reviewIssues>0",
            "priority": "P1",
            "problem": "{name} 初审问题数 {value}",
            "action": "追溯源头，优化需求质量",
        },
        {
            "condition": "invalidData>0",
            "priority": "P1",
            "problem": "{name} 无效数据 {value} 条",
            "action": "排查无效数据原因，提升数据质量",
        },
    ],
    "monthlyTrend": monthly_trend_dict,
    # 新模板扩展数据
    "qualityReview": quality_review,
    "preReview": preview_stats,
    "issueFeedback": issue_feedback if issue_feedback else None,
    "productAcceptance": product_acceptance if product_acceptance.get("by_month") else None,
    "collectionCapacity": collection_capacity,
}

# 写入文件
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# 同步内嵌数据到 index.html 的 dataEmbed 块（离线 fallback 保持新鲜，避免显示过时快照）
# 仅在输出为默认 data.json 时同步；跑 test.json 等自定义输出不碰 index.html
if os.path.basename(OUTPUT_PATH) == _default_out:
    sync_embed_data(os.path.join(os.path.dirname(__file__), 'index.html'), data)

print(f"✅ 解析完成！共 {len(tasks)} 条任务, {len(settlements)} 条结算记录")
print(f"   月份: {months_in_data}")
print(f"   需求项: {demand_items}")
print(f"   结算类型: {settle_types}")
print(f"   问题反馈: {issue_feedback.get('问题反馈量', 'N/A') if issue_feedback else '无'}")
if preview_stats:
    print(f"   加工预审: {preview_stats['total']} 条, 通过率 {preview_stats.get('passRate', 0)}%, 新供应商 {preview_stats['new_enterprise']}, 老供应商 {preview_stats['old_enterprise']}, 价格审核 {preview_stats['price_audit']}")
print(f"   采集产能: {len(collection_capacity)} 条")
print(f"   台账合并: {_merged_count} 条")
print(f"   已写入: {OUTPUT_PATH}")

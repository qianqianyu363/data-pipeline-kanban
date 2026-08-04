#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
升级 看板数据录入模板-调整.xlsx：
1. 各项需求跟踪 第7列插入「约定采集完成时间」
2. 预算与结算 第7-11列插入去年值5列
3. 加工预审 第9列插入「价格审核结果」
4. 新建「采集产能」Sheet
插入后自动修正所有公式引用的列号（+偏移量）。
"""
import openpyxl, re, shutil

SRC = '看板数据录入模板-调整.xlsx'
OUT = '看板数据录入模板-调整_升级版.xlsx'
BAK = '.workbuddy/backup_20260803/看板数据录入模板-调整_HEAD.xlsx'

# 备份当前(HEAD)模板
shutil.copy(SRC, BAK)

wb = openpyxl.load_workbook(SRC)  # data_only=False 保留公式


def letters_to_num(letters):
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def num_to_letters(n):
    res = ''
    while n:
        n, r = divmod(n - 1, 26)
        res = chr(65 + r) + res
    return res


def shift_formula(f, insert_col, offset):
    """公式中所有列号>insert_col的引用 +offset，兼容 $ 绝对引用"""
    def repl(m):
        prefix = m.group(1)          # 可能含 $
        letters = prefix.replace('$', '')
        col_num = letters_to_num(letters)
        if col_num > insert_col:
            col_num += offset
        new_letters = num_to_letters(col_num)
        if '$' in prefix:
            new_letters = '$' + new_letters
        return new_letters + m.group(2)
    return re.sub(r'(\$?[A-Za-z]{1,3}\$?)(\d+)', repl, f)


def insert_and_shift(ws, insert_col, amount):
    ws.insert_cols(insert_col, amount)
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = shift_formula(cell.value, insert_col, amount)


# ---- 1. 各项需求跟踪: 插入第7列 约定采集完成时间 ----
ws = wb['各项需求跟踪']
insert_and_shift(ws, 7, 1)
ws.cell(row=2, column=7, value='约定采集完成时间')

# ---- 2. 预算与结算: 插入第7-11列 去年值5列 ----
ws = wb['预算与结算']
insert_and_shift(ws, 7, 5)
for i, h in enumerate(['去年产品模块', '去年需求类型', '去年需求数量', '去年结算单价', '去年支出费用']):
    ws.cell(row=2, column=7 + i, value=h)

# ---- 修复预算与结算的合并单元格（openpyxl 不移动合并区，需手动右移5列）----
from openpyxl.utils import get_column_letter
merges = list(ws.merged_cells.ranges)
for mr in merges:
    ws.unmerge_cells(str(mr))
ws.cell(row=2, column=16, value='结算一览')   # 原K2标题，插入后应在P2
for mr in merges:
    if mr.min_col >= 7:
        new_range = '%s%d:%s%d' % (get_column_letter(mr.min_col + 5), mr.min_row,
                                   get_column_letter(mr.max_col + 5), mr.max_row)
        ws.merge_cells(new_range)
    else:
        ws.merge_cells(str(mr))

# ---- 3. 加工预审: 插入第9列 价格审核结果（表头在第1行）----
ws = wb['加工预审']
insert_and_shift(ws, 9, 1)
ws.cell(row=1, column=9, value='价格审核结果')

# ---- 4. 新建「采集产能」Sheet ----
if '采集产能' not in wb.sheetnames:
    ws = wb.create_sheet('采集产能')
else:
    ws = wb['采集产能']
for i, h in enumerate(['月份', '计划采集家数', '计划采集条数', '实际采集家数', '实际采集条数', '员工总人数', '备注']):
    ws.cell(row=1, column=1 + i, value=h)

wb.save(OUT)
print('OK, saved to', OUT, 'Sheets:', wb.sheetnames)
